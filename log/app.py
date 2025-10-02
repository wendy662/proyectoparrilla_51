# app.py unificado (comentado por bloques y funciones)
from flask import (
    Flask, render_template, request, redirect, url_for, session,
    jsonify, flash, send_file
)
from flask_mysqldb import MySQL
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash


# from functools import wraps
import uuid
import io

# para la fecha de la reserva
from datetime import datetime
# Reportes
from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

# ------------------ CONFIG ------------------
# Aquí se configura la app Flask, la conexión a MySQL y el servicio de correo.
# También se define la secret_key para sesiones.
app = Flask(__name__, template_folder="templates")

app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'parrilla51'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

mysql = MySQL(app)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'andresfariasa@juandelcorral.edu.co'
app.config['MAIL_PASSWORD'] = 'ekyxnfjsubefudgm'
mail = Mail(app)

app.secret_key = "pinchellave"

# ------------------ DECORADOR LOGIN ------------------
# Decorador para proteger rutas que requieren autenticación.
# Si no hay sesión con 'logueado', redirige al login y muestra aviso.

# ------------------ AUTENTICACIÓN (ruta raíz) ------------------
# La misma vista maneja '/' y '/login' (registrada también como
# endpoint 'login').
# - GET: muestra template index.html
# - POST: procesa login (busca usuario activo y valida contraseña)


@app.route('/', methods=['GET', 'POST'])
def index():
    # Manejo login si se hace POST (compatible con formularios
    # que hagan action a '/')
    if request.method == 'POST':
        correo = request.form.get('txtCorreo')
        password = request.form.get('txtPassword')
        cur = mysql.connection.cursor()
        cur.execute('SELECT * FROM usuarios WHERE correo=%s '
                    'AND estado="activo"', (correo,))
        account = cur.fetchone()
        if account and check_password_hash(account.get('contraseña', ''),
                                           password):
            # Crear sesión y guardar datos importantes
            session['logueado'] = True
            session['id_usuario'] = account['id_usuario']
            session['nombre'] = account['nombre']
            session['rol'] = account.get('rol', 'cliente')
            # redirigir según rol
            if session['rol'] == 'administrador':
                return redirect(url_for("admin_dashboard"))
            elif session['rol'] == 'cliente':
                return redirect(url_for("cliente_dashboard"))
            elif session['rol'] == 'empleado':
                return redirect(url_for("empleado_dashboard"))
        else:
            # Mensaje sencillo si falla autenticación o cuenta no está activa
            return render_template('index.html', mensaje="Usuario o contraseña"
                                   "incorrectos, o cuenta no activada")
    return render_template('index.html')

# Registrar alias '/login' con endpoint 'login' apuntando a la misma vista
# index


# app.add_url_rule('/login', endpoint='login', view_func=index,
#               methods=['GET', 'POST'])

# Ruta para mostrar formulario de registro


@app.route('/registro')
def registro():
    return render_template('registro.html')

# Guardar usuario: recibe datos del formulario, encripta contraseña, crea token
# y envía correo de activación con enlace que apunta a /activar/<token>


@app.route('/guardar-usuario', methods=["POST"])
def guardar_usuario():
    nombre = request.form['nombre']
    apellido = request.form['apellido']
    telefono = request.form['telefono']
    direccion = request.form['direccion']
    correo = request.form['correo']
    password = request.form['password']
    rol = request.form.get('rol', 'cliente')

    password_hash = generate_password_hash(password)
    token = str(uuid.uuid4())

    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO usuarios(nombre, apellido, telefono, direccion, correo,
        contraseña, rol, estado, token_activacion)
        VALUES(%s,%s,%s,%s,%s,%s,%s,'inactivo',%s)
    """, (nombre, apellido, telefono, direccion, correo,
          password_hash, rol, token))
    mysql.connection.commit()

    enlace = url_for('activar_cuenta', token=token, _external=True)
    try:
        msg = Message('Activa tu cuenta', sender=app.config['MAIL_USERNAME'],
                      recipients=[correo])
        msg.body = f'Hola {nombre}, haz clic en el siguiente enlace para'
        f'activar tu cuenta:\n{enlace}'
        mail.send(msg)
    except Exception as e:
        # Loguea error de envío de correo en consola; no romper la experiencia
        # de registro
        print("Error enviando correo:", e)

    return render_template('index.html',
                           mensaje="Revisa tu correo para activar tu cuenta")

# Activación de cuenta: recibe token, busca usuario y cambia estado a 'activo'


@app.route('/activar/<token>')
def activar_cuenta(token):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM usuarios WHERE token_activacion=%s", (token,))
    usuario = cur.fetchone()
    if usuario:
        cur.execute("UPDATE usuarios SET estado='activo'",
                    "token_activacion=NULL WHERE id_usuario=%s",
                    (usuario['id_usuario'],))
        mysql.connection.commit()
        return ("Tu cuenta ha sido activada correctamente",
                "Ya puedes iniciar sesión.")

    else:
        return "Token inválido o expirado."

# Ruta histórica para compatibilidad con templates antiguos
# Redirige al index que ya maneja el POST de login


# @app.route('/acceso-login', methods=["POST"])
# def acceso_login():
#    return index()

# ------------------ DASHBOARDS ------------------
# Vistas protegidas por login_requerido. Cada una muestra el template
# correspondiente.


@app.route("/admin")
def admin_dashboard():
    return render_template("admin2.html")


@app.route("/cliente/dashboard")
def cliente_dashboard():
    return render_template("cliente_dashboard.html")


@app.route("/empleado")
def empleado_dashboard():
    return render_template("empleado.html")

# ------------------ PERFIL ------------------
# GET: devuelve datos del perfil del usuario logueado en formato JSON
# POST: actualiza el perfil del usuario (recibe JSON con nuevos datos)


@app.route('/perfil', methods=['GET'])
def perfil():
    cur = mysql.connection.cursor()
    cur.execute("SELECT nombre, apellido, telefono, direccion,"
                "correo FROM usuarios WHERE id_usuario=%s",
                (session['id_usuario'],))
    usuario = cur.fetchone()
    return jsonify(usuario or {})


@app.route('/perfil', methods=['POST'])
def actualizar_perfil():
    data = request.json
    cur = mysql.connection.cursor()
    cur.execute("""
        UPDATE usuarios
        SET nombre=%s, apellido=%s, telefono=%s, direccion=%s, correo=%s
        WHERE id_usuario=%s
    """, (data['nombre'], data['apellido'], data['telefono'],
          data['direccion'],
          data['correo'], session['id_usuario']))
    mysql.connection.commit()
    return jsonify({'mensaje': 'Perfil actualizado correctamente'})

# ------------------ BLUEPRINT-LIKE ROUTES (integradas) ------------------
# Implementación de rutas que actúan como si fueran blueprints
# integrados en un único archivo.
# Se crean alias para endpoints que templates antiguos
# podrían usar (evita errores con url_for).
# Reservas (vista sencilla que lista todas las reservas)


@app.route("/cliente_reservar", methods=["GET", "POST"])
def cliente_reservar():
    cur = mysql.connection.cursor()

    if request.method == "POST":
        fecha = request.form["fecha"]
        hora = request.form["hora"]
        cant_personas = request.form["cant_personas"]
        id_mesa = request.form["id_mesa"]
        telefono = request.form["telefono"]

        # Guardar en MySQL
        cur.execute("""
            INSERT INTO reservas (fecha, hora, cant_personas,
            id_mesa, telefono)
            VALUES (%s, %s, %s, %s, %s)
        """, (fecha, hora, cant_personas, id_mesa, telefono))
        mysql.connection.commit()

        flash("✅ Reserva creada con éxito", "success")
        return redirect(url_for("ver_reservas"))  # Te redirige a ver las
    # creservas

    # Si es GET: cargar mesas disponibles para el select
    cur.execute("SELECT id_mesa, numero, capacidad FROM mesas")
    mesas = cur.fetchall()

    return render_template("cliente_reservar.html", mesas=mesas)

# Mesas: listar mesas ordenadas por número


@app.route("/mesas")
def ver_mesas():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM mesas ORDER BY numero ASC")
    mesas = cur.fetchall()
    return render_template("mesas.html", mesas=mesas)

# Alias para templates que esperan otro endpoint


app.add_url_rule('/mesas/ver', endpoint='mesas.mesas',
                 view_func=ver_mesas, methods=['GET'])

# ------------------ FUNCIONES AUX: CATEGORIAS ------------------
# Función auxiliar para obtener sólo ciertas categorías importantes


def obtener_categorias():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT * FROM categorias
        WHERE nombre_categoria IN ('Bebidas', 'Res', 'Pollo',
        'Cerdo', 'Plato del dia',
        'Entradas', 'Acompañamientos', 'Platos combinados',
        'Cortes gruesos','Adicionales')
    """)
    return cur.fetchall()

# ------------------ INVENTARIO y CRUD ------------------
# Vista inicial admin


@app.route("/inicioadmin")
def inicioadmin():
    return render_template("inicioadmin.html")

# Inventario: reúne productos, insumos, mesas y subcategorías para la vista


@app.route("/inventario")
def inventario():
    cur = mysql.connection.cursor()
    # Productos con categoría
    # (LEFT JOIN permite que no haya categoría sin fallar)
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio,
               p.imagen, c.nombre_categoria, p.cod_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
    """)
    productos = cur.fetchall()

    # Insumos con su subcategoría
    cur.execute("""
        SELECT i.id_insumo, i.nombre, i.cantidad, i.precio,
        i.fecha_vencimiento, i.lote,
               s.nombre_subcategoria, i.subcategoria_id
        FROM insumos i
        LEFT JOIN subcategorias_insumos s ON
        i.subcategoria_id = s.id_subcategoria
    """)
    insumos = cur.fetchall()

    # Mesas
    cur.execute("SELECT * FROM mesas")
    mesas = cur.fetchall()

    categorias = obtener_categorias()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()

    return render_template("inventario.html",
                           productos=productos,
                           insumos=insumos,
                           mesas=mesas,
                           categorias=categorias,
                           subcategorias=subcategorias)

# CRUD Productos: agregar producto
# (GET muestra formulario, POST procesa y guarda)


@app.route("/producto/agregar", methods=["GET", "POST"])
def agregar_producto():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            descripcion = request.form["descripcion"].strip()
            # Una sola línea, limpio y correcto
            precio = int(request.form
                         ["precio"].replace(".", "").replace(",", ""))

            cod_categoria = int(request.form["cod_categoria"])
            imagen = request.form.get("imagen", "").strip()
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO productos (nombre, cantidad, descripcion, precio,
            cod_categoria, imagen)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    categorias = obtener_categorias()
    return render_template("editar_producto.html",
                           producto=None, categorias=categorias)

# Editar producto: GET carga producto; POST actualiza


@app.route("/producto/editar/<int:id_producto>", methods=["GET", "POST"])
def editar_producto(id_producto):
    cur = mysql.connection.cursor()
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            descripcion = request.form["descripcion"].strip()
            precio = int(request.form
                         ["precio"].replace(".", "").replace(",", ""))
            cod_categoria = int(request.form["cod_categoria"])
            imagen = request.form.get("imagen", "").strip()
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur.execute("""
            UPDATE productos
            SET nombre=%s, cantidad=%s, descripcion=%s, precio=%s,
            cod_categoria=%s, imagen=%s
            WHERE id_producto=%s
        """, (nombre, cantidad, descripcion, precio, cod_categoria, imagen,
              id_producto))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur.execute("SELECT * FROM productos WHERE id_producto=%s", (id_producto,))
    producto = cur.fetchone()
    categorias = obtener_categorias()
    if not producto:
        return "Producto no encontrado", 404
    return render_template("editar_producto.html", producto=producto,
                           categorias=categorias)

# Eliminar producto (POST para evitar CSRF por GET)


@app.route("/producto/eliminar/<int:id_producto>", methods=["POST"])
def eliminar_producto(id_producto):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM productos WHERE id_producto=%s", (id_producto,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# CRUD Insumos: estructuras paralelas a productos (agregar, editar, eliminar)


@app.route("/insumo/agregar", methods=["GET", "POST"])
def agregar_insumo():
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            precio = float(request.form["precio"])
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None
            lote = request.form.get("lote") or None
            subcategoria_id = int(request.form["subcategoria_id"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO insumos (nombre, cantidad, precio, fecha_vencimiento,
            lote, subcategoria_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (nombre, cantidad, precio, fecha_vencimiento, lote,
              subcategoria_id))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()
    return render_template("editar_insumo.html", insumo=None,
                           subcategorias=subcategorias)


@app.route("/insumo/editar/<int:id_insumo>", methods=["GET", "POST"])
def editar_insumo(id_insumo):
    cur = mysql.connection.cursor()
    if request.method == "POST":
        try:
            nombre = request.form["nombre"].strip()
            cantidad = int(request.form["cantidad"])
            precio = float(request.form["precio"])
            fecha_vencimiento = request.form.get("fecha_vencimiento") or None
            lote = request.form.get("lote") or None
            subcategoria_id = int(request.form["subcategoria_id"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400

        cur.execute("""
            UPDATE insumos
            SET nombre=%s, cantidad=%s, precio=%s, fecha_vencimiento=%s,
            lote=%s, subcategoria_id=%s
            WHERE id_insumo=%s
        """, (nombre, cantidad, precio, fecha_vencimiento, lote,
              subcategoria_id, id_insumo))
        mysql.connection.commit()
        return redirect(url_for("inventario"))

    cur.execute("SELECT * FROM insumos WHERE id_insumo=%s", (id_insumo,))
    insumo = cur.fetchone()
    cur.execute("SELECT * FROM subcategorias_insumos")
    subcategorias = cur.fetchall()
    if not insumo:
        return "Insumo no encontrado", 404
    return render_template("editar_insumo.html", insumo=insumo,
                           subcategorias=subcategorias)


@app.route("/insumo/eliminar/<int:id_insumo>", methods=["POST"])
def eliminar_insumo(id_insumo):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM insumos WHERE id_insumo=%s", (id_insumo,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# ------------------ CRUD Mesas ------------------
# Agregar mesa, cambiar estado y eliminar


@app.route("/mesa/agregar", methods=["GET", "POST"])
def agregar_mesa():
    if request.method == "POST":
        try:
            numero = int(request.form["numero"])
            capacidad = int(request.form["capacidad"])
        except (ValueError, KeyError):
            return "Datos inválidos", 400
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO mesas (numero, capacidad, estado)
            VALUES (%s, %s, 'Disponible')
        """, (numero, capacidad))
        mysql.connection.commit()
        return redirect(url_for("inventario"))
    return render_template("editar_mesa.html", mesa=None)


@app.route("/mesa/cambiar_estado/<int:id_mesa>")
def cambiar_estado(id_mesa):
    cur = mysql.connection.cursor()
    cur.execute("SELECT estado FROM mesas WHERE id_mesa=%s", (id_mesa,))
    mesa = cur.fetchone()
    if not mesa:
        return "Mesa no encontrada", 404
    nuevo_estado = (
        "Disponible" if mesa["estado"] == "No disponible" else "No disponible"
        )
    cur.execute("UPDATE mesas SET estado=%s WHERE id_mesa=%s",
                (nuevo_estado, id_mesa))
    mysql.connection.commit()
    return redirect(url_for("inventario"))


@app.route("/mesa/eliminar/<int:id_mesa>", methods=["POST"])
def eliminar_mesa(id_mesa):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM mesas WHERE id_mesa=%s", (id_mesa,))
    mysql.connection.commit()
    return redirect(url_for("inventario"))

# Confirmación de eliminación: prepara URL para el POST real de eliminación


@app.route("/confirmar_eliminacion/<string:tipo>/<int:item_id>")
def confirmar_eliminacion(tipo, item_id):
    if tipo == "producto":
        url_elim = url_for("eliminar_producto", id_producto=item_id)
    elif tipo == "insumo":
        url_elim = url_for("eliminar_insumo", id_insumo=item_id)
    elif tipo == "mesa":
        url_elim = url_for("eliminar_mesa", id_mesa=item_id)
    else:
        return "Tipo no válido", 400
    return render_template("eliminar.html", tipo=tipo,
                           item_id=item_id, url=url_elim)

# ------------------ ASIGNAR ROL / USUARIOS ------------------
# Vista para listar usuarios y asignar rol/estado via formulario POST


@app.route("/asignarol", methods=["GET", "POST"])
def asignarol():
    if request.method == "POST":
        user_id = request.form["id_usuario"]
        nuevo_rol = request.form["rol"]
        nuevo_estado = request.form["estado"]
        cur = mysql.connection.cursor()
        cur.execute("""
            UPDATE usuarios
            SET rol = %s, estado = %s
            WHERE id_usuario = %s
        """, (nuevo_rol, nuevo_estado, user_id))
        mysql.connection.commit()
        return redirect(url_for("asignarol"))
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM usuarios")
    usuarios = cur.fetchall()
    return render_template("asignarol.html", usuarios=usuarios)

# Rutas helper para cambiar estado/rol vía URL
# (útiles para botones rápidos en la UI)


@app.route("/usuario/cambiar_estado/<int:id_usuario>/<string:nuevo_estado>")
def cambiar_estado_usuario(id_usuario, nuevo_estado):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE usuarios SET estado = %s WHERE id_usuario = %s",
                (nuevo_estado, id_usuario))
    mysql.connection.commit()
    return redirect(url_for("asignarol"))


@app.route("/usuario/cambiar_rol/<int:id_usuario>/<string:nuevo_rol>")
def cambiar_rol_usuario(id_usuario, nuevo_rol):
    cur = mysql.connection.cursor()
    cur.execute("UPDATE usuarios SET rol = %s WHERE id_usuario = %s",
                (nuevo_rol, id_usuario))
    mysql.connection.commit()
    return redirect(url_for("asignarol"))

# ------------------ CONSULTAS Y REPORTES ------------------
# Consultas específicas para reservas y ventas, con filtros por estado/tipo


@app.route("/consultar_reservas")
def consultar_reservas():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id_reserva, fecha, hora, cant_personas, estado, cod_mesa,
        telefono, id_usuario
        FROM reservas
        WHERE estado IN ('aceptada', 'cancelada')
        ORDER BY fecha ASC, hora ASC
    """)
    reservas = cur.fetchall()
    return render_template("consultar_reservas.html", reservas=reservas)


@app.route("/consultaVentas")
def consultaVentas():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora,
        p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante' AND p.estado IN ('entregado',
        'cancelado')
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio' AND p.estado IN ('entregado',
        'cancelado')
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos_domicilio = cur.fetchall()

    return render_template("consultaVentas.html",
                           pedidos_restaurante=pedidos_restaurante,
                           pedidos_domicilio=pedidos_domicilio)

# Consultas de inventario para vistas específicas


@app.route("/consulta_P")
def consulta_P():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.cantidad, p.descripcion, p.precio,
               p.imagen, c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.nombre ASC
    """)
    productos = cur.fetchall()
    return render_template("consulta_P.html", productos=productos)


@app.route("/consulta_Y")
def consulta_Y():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT i.id_insumo, i.nombre, i.cantidad, i.precio,
        i.fecha_vencimiento, i.lote,
               s.nombre_subcategoria
        FROM insumos i
        LEFT JOIN subcategorias_insumos s ON
        i.subcategoria_id = s.id_subcategoria
        ORDER BY i.nombre ASC
    """)
    insumos = cur.fetchall()
    return render_template("consulta_Y.html", insumos=insumos)

# ------------------ REPORTES (pantalla / pdf / excel) ------------------
# Vista que consolida distintos datos para el módulo de reportes


@app.route("/reportes")
def reportes():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall() or []

    cur.execute("SELECT * FROM vista_insumos_stock_bajo")
    stock_bajo = cur.fetchall() or []

    return render_template("reportes.html",
                           pedidos_restaurante=pedidos_restaurante or [],
                           pedidos_domicilio=pedidos_domicilio or [],
                           reservas=reservas,
                           stock_bajo=stock_bajo)

# Genera PDF con reportes: pedidos restaurante, pedidos domicilio,
# reservas e inventario bajo


@app.route("/reportes/pdf")
def reportes_pdf():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall()

    cur.execute("SELECT * FROM insumos WHERE cantidad < 5")
    inventario = cur.fetchall()

    # Construcción de PDF usando reportlab
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Pedidos restaurante
    elements.append(Paragraph("Pedidos en Restaurante", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                         str(p['fecha']), str(p['hora']),
                         f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Pedidos domicilio
    elements.append(Paragraph("Pedidos a Domicilio", styles['Heading2']))
    data = [["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"]]
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            data.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                         str(p['fecha']), str(p['hora']),
                         f"${p['total']:.2f}", p['estado']])
    else:
        data.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Reservas
    elements.append(Paragraph("Reservas", styles['Heading2']))
    data = [["ID", "Fecha", "Hora", "Personas", "Estado", "Mesa", "Capacidad"]]
    if reservas:
        for r in reservas:
            data.append([r['id_reserva'], str(r['fecha']), str(r['hora']),
                         r['cant_personas'], r['estado'], r.get('mesa'),
                         r.get('capacidad')])
    else:
        data.append(["-", "No hay reservas registradas", "-", "-",
                     "-", "-", "-"])
    elements.append(Table(data))
    elements.append(Spacer(1, 12))

    # Inventario bajo
    elements.append(Paragraph("Inventario Bajo", styles['Heading2']))
    data = [["ID Insumo", "Nombre", "Cantidad", "Precio"]]
    if inventario:
        for i in inventario:
            data.append([i['id_insumo'], i['nombre'], i['cantidad'],
                         f"${i['precio']:.2f}"])
    else:
        data.append(["-", "No hay insumos con stock bajo", "-", "-"])
    elements.append(Table(data))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name="reporte_general.pdf",
                     mimetype="application/pdf")

# Genera Excel con reportes análogos al PDF


@app.route("/reportes/excel")
def reportes_excel():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha,
        p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'restaurante'
    """)
    pedidos_restaurante = cur.fetchall()

    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido,
        p.fecha, p.hora, p.total, p.estado
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        WHERE p.tipo_entrega = 'domicilio'
    """)
    pedidos_domicilio = cur.fetchall()

    cur.execute("SELECT * FROM vista_reservas_mesas")
    reservas = cur.fetchall()

    cur.execute("SELECT * FROM insumos WHERE cantidad < 5")
    inventario = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos Restaurante"
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_restaurante:
        for p in pedidos_restaurante:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                       str(p['fecha']), str(p['hora']),
                       p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos en restaurante", "-", "-", "-", "-"])

    ws = wb.create_sheet("Pedidos Domicilio")
    ws.append(["ID", "Cliente", "Fecha", "Hora", "Total", "Estado"])
    if pedidos_domicilio:
        for p in pedidos_domicilio:
            ws.append([p['id_pedido'], f"{p['nombre']} {p['apellido']}",
                       str(p['fecha']), str(p['hora']),
                       p['total'], p['estado']])
    else:
        ws.append(["-", "No hay pedidos a domicilio", "-", "-", "-", "-"])

    ws = wb.create_sheet("Reservas")
    ws.append(["ID Reserva", "Fecha", "Hora", "Personas",
               "Estado", "Mesa", "Capacidad"])
    if reservas:
        for r in reservas:
            ws.append([r['id_reserva'], str(r['fecha']),
                       str(r['hora']), r['cant_personas'], r['estado'],
                       r.get('mesa'), r.get('capacidad')])
    else:
        ws.append(["-", "No hay reservas registradas", "-",
                   "-", "-", "-", "-"])

    ws = wb.create_sheet("Inventario Bajo")
    ws.append(["ID Insumo", "Nombre", "Cantidad", "Precio"])
    if inventario:
        for i in inventario:
            ws.append([i['id_insumo'], i['nombre'],
                       i['cantidad'], i['precio']])
    else:
        ws.append(["-", "No hay insumos con stock bajo", "-", "-"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer, as_attachment=True,
        download_name="reporte_general.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet")

# ------------------ ADMIN: PEDIDOS ------------------
# Vista de pedidos para administrador y función para cambiar estado


@app.route("/admin/pedidos")
def admin_pedidos():
    if session.get("rol") != "administrador":
        flash("Acceso denegado", "danger")
        return redirect(url_for("index"))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_pedido, u.nombre, u.apellido, p.fecha, p.hora,
        p.total, p.estado, p.tipo_entrega, p.cod_mesa
        FROM pedidos p
        LEFT JOIN usuarios u ON p.cod_usuario = u.id_usuario
        ORDER BY p.fecha DESC, p.hora DESC
    """)
    pedidos = cur.fetchall()

    return render_template("admin_pedidos.html", pedidos=pedidos)


@app.route(
    "/admin/pedidos/cambiar_estado/<int:id_pedido>/<string:nuevo_estado>"
    )
def cambiar_estado_pedido(id_pedido, nuevo_estado):
    if session.get("rol") != "administrador":
        flash("Acceso denegado", "danger")
        return redirect(url_for("index"))

    cur = mysql.connection.cursor()
    cur.execute("UPDATE pedidos SET estado=%s WHERE id_pedido=%s",
                (nuevo_estado, id_pedido))
    mysql.connection.commit()
    flash(f"Estado del pedido {id_pedido} cambiado a {nuevo_estado}",
          "success")
    return redirect(url_for("admin_pedidos"))

# ------------------ RUTAS CLIENTE ------------------
# Cliente puede reservar, ver productos y usar carrito/pedidos


@app.route("/cliente/reservar/form", methods=["GET", "POST"])
def cliente_reservar_form():
    if request.method == "POST":
        nombre = request.form["nombre"]
        documento = request.form["documento"]
        fecha = request.form["fecha"]
        hora = request.form["hora"]
        cant_personas = int(request.form["cant_personas"])
        tipo_evento = request.form["tipo_evento"]
        comentarios = request.form["comentarios"]
        id_mesa = int(request.form["id_mesa"])
        telefono = request.form["telefono"]

        #  Validar que la fecha no sea pasada
        fecha_reserva = datetime.strptime(fecha, "%Y-%m-%d").date()
        hoy = datetime.now().date()

        if fecha_reserva < hoy:
            flash("❌ No puedes reservar en una fecha que ya pasó.", "error")
            return redirect(url_for("cliente_reservar_form"))

        #  Validar que si la fecha es hoy, la hora no haya pasado
        if fecha_reserva == hoy:
            hora_actual = datetime.now().time()
            hora_reserva = datetime.strptime(hora, "%H:%M").time()
            if hora_reserva < hora_actual:
                flash("❌ No puedes reservar en una hora que ya pasó.", "error")
                return redirect(url_for("cliente_reservar_form"))

        # ✅ Si pasa la validación, guardar la reserva en la base de datos
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO reservas (nombre, documento, fecha, hora,
            cant_personas, tipo_evento, comentarios, estado, cod_mesa,
            telefono, id_usuario)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'pendiente', %s, %s, %s)
        """, (
            nombre,
            documento,
            fecha,
            hora,
            cant_personas,
            tipo_evento,
            comentarios,
            id_mesa,
            telefono,
            1
        ))

        print("🔎 CONSULTA EJECUTADA:")

        mysql.connection.commit()

        flash("✅ Reserva registrada correctamente. Espera confirmación.",
              "success")
        return redirect(url_for("cliente_dashboard"))

    # Si el método es GET → mostrar mesas disponibles
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM mesas WHERE estado='Disponible'")
    mesas = cur.fetchall()

    return render_template("cliente_reservar.html", mesas=mesas)


@app.route('/cancelar/<int:id>', methods=['POST'])
def cancelar_reserva(id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM reservas WHERE id_reserva = %s", (id,))
    mysql.connection.commit()
    cur.close()
    flash("Reserva cancelada exitosamente ✅", "success")
    return redirect(url_for('ver_reservas'))


@app.route("/mis-reservas", methods=["GET"])
def ver_reservas():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM reservas")
    reservas = cur.fetchall()

    return render_template("cliente_ver_reservas.html", reservas=reservas)


# Mostrar productos disponibles al cliente


@app.route("/cliente/productos")
def cliente_productos():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.id_producto, p.nombre, p.descripcion, p.precio,
        p.imagen, c.nombre_categoria
        FROM productos p
        LEFT JOIN categorias c ON p.cod_categoria = c.id_categoria
        ORDER BY p.nombre ASC
    """)
    productos = cur.fetchall()
    return render_template("cliente_productos.html", productos=productos)

# Agregar producto al carrito guardado en sesión


@app.route("/carrito/agregar/<int:id_producto>", methods=["POST"])
def agregar_carrito(id_producto):
    cantidad = int(request.form.get("cantidad", 1))
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM productos WHERE id_producto=%s", (id_producto,))
    producto = cur.fetchone()
    if not producto:
        return "Producto no encontrado", 404

    if "carrito" not in session:
        session["carrito"] = []

    carrito = session["carrito"]
    for item in carrito:
        if item["id_producto"] == id_producto:
            item["cantidad"] += cantidad
            break
    else:
        carrito.append({
            "id_producto": id_producto,
            "nombre": producto["nombre"],
            "precio": producto["precio"],
            "cantidad": cantidad
        })

    session["carrito"] = carrito
    flash("Producto agregado al carrito", "success")
    return redirect(url_for("cliente_productos"))

# Mostrar carrito (incluye total y mesas disponibles
# para seleccionar si aplica)


@app.route("/carrito", endpoint="cliente_carrito")
def ver_carrito():
    carrito = session.get("carrito", [])
    total = sum(item["precio"] * item["cantidad"] for item in carrito)

    # ✅ Traer mesas disponibles de la base de datos
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM mesas WHERE estado='Disponible'")
    mesas = cur.fetchall()

    return render_template("cliente_carrito.html", carrito=carrito,
                           total=total, mesas=mesas)

# Confirmar pedido: crea pedido, detalle y limpia carrito de la sesión


@app.route("/pedido/confirmar", methods=["POST"])
def hacer_pedido():
    carrito = session.get("carrito", [])
    if not carrito:
        flash("El carrito está vacío", "warning")
        return redirect(url_for("cliente_productos"))

    total = sum(item["precio"] * item["cantidad"] for item in carrito)

    tipo_entrega = request.form.get("tipo_entrega", "mesa")
    metodo_pago = request.form.get("metodo_pago", "efectivo")
    cod_mesa = request.form.get("cod_mesa") if tipo_entrega == "mesa" else None

    id_usuario = session.get("id_usuario", 1)  # para hacer la prueba con un id

    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO pedidos (cod_usuario, fecha, hora, total, estado,
        tipo_entrega, metodo_pago, cod_mesa)
        VALUES (%s, CURDATE(), CURTIME(), %s, 'pendiente', %s, %s, %s)
    """, (id_usuario, total, tipo_entrega, metodo_pago, cod_mesa))
    id_pedido = cur.lastrowid

    for item in carrito:
        cur.execute("""
            INSERT INTO detalle_pedido (cod_pedido, cod_producto,
            cantidad, precio_unitario)
            VALUES (%s, %s, %s, %s)
        """, (id_pedido, item["id_producto"],
              item["cantidad"], item["precio"]))

    mysql.connection.commit()
    session.pop("carrito", None)

    flash("Pedido registrado correctamente", "success")
    return redirect(url_for("cliente_dashboard"))


@app.route("/mis_pedidos")
def ver_pedidos():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id_pedido, cod_usuario, fecha, hora, total, estado,
        tipo_entrega, metodo_pago, cod_mesa
        FROM pedidos ORDER BY id_pedido DESC
    """)
    pedidos = cur.fetchall()
    return render_template("mis_pedidos.html", pedidos=pedidos)


# ✅ ELIMINAR PRODUCTO DEL CARRITO


@app.route("/carrito/eliminar/<int:id_producto>")
def eliminar_carrito(id_producto):
    carrito = session.get("carrito", [])
    nuevo_carrito = [
        item for item in carrito if item["id_producto"] != id_producto]
    session["carrito"] = nuevo_carrito
    flash("Producto eliminado del carrito", "success")
    return redirect(url_for("cliente_carrito"))

# ------------------ LOGOUT ------------------
# Limpia la sesión completamente


# @app.route("/logout")
# def logout():
#    session.clear()
#    return redirect(url_for("login"))

# ------------------ RUN ------------------
# Ejecuta la app en modo debug (útil durante desarrollo)


#--------------------EMPLEADOOOOO----------------------

@app.route("/testdb")
def testdb():
    try:
        cur = mysql.connection.cursor()
        cur.execute("SHOW TABLES;")   # consulta las tablas de tu BD
        tables = cur.fetchall()
        cur.close()
        return f"✅ Conectado. Tablas en la BD: {tables}"
    except Exception as e:
        return f"❌ Error conectando a MySQL: {str(e)}"

# ================== RUTAS EMPLEADO ==================
from flask import Flask, render_template, request, redirect, url_for
import mysql.connector

app = Flask(__name__)

# 📌 Conexión a la BD
conexion = mysql.connector.connect(
    host="localhost",
    user="root",          # <-- cambia si tu usuario es distinto
    password="",          # <-- tu clave aquí
    database="parrilla51"
)
cursor = conexion.cursor(dictionary=True)  # <-- devuelve diccionarios {columna: valor}
from flask import Flask, render_template, request, redirect, url_for
import mysql.connector

app = Flask(__name__)
from flask import Flask, render_template, request, redirect, url_for
import mysql.connector

app = Flask(__name__)

from flask import Flask, render_template, request, redirect, url_for
import mysql.connector

app = Flask(__name__)
from flask import Flask, render_template, request, redirect, url_for
import mysql.connector

from flask import Flask, render_template, request, redirect, url_for, jsonify
import mysql.connector

app = Flask(__name__)

# ===================== CONEXIÓN =====================
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="parrilla51"
    )


# ===================== LISTAR PRODUCTOS Y CATEGORÍAS =====================
@app.route("/registrar_empleado")
def registrar_empleado():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id_categoria, nombre_categoria FROM categorias")
    categorias = cursor.fetchall()

    cursor.execute("""
        SELECT p.id_producto, p.nombre, p.precio, p.descripcion, c.nombre_categoria
        FROM productos_empleados p
        JOIN categorias c ON p.id_categoria = c.id_categoria
    """)
    productos = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("registrar_empleado.html", productos=productos, categorias=categorias)


# ===================== AGREGAR CATEGORÍA =====================
@app.route("/agregar_categoria", methods=["POST"])
def agregar_categoria():
    nombre_categoria = request.form["nombre_categoria"]

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO categorias (nombre_categoria) VALUES (%s)", (nombre_categoria,))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("registrar_empleado"))


# ===================== EDITAR CATEGORÍA =====================
@app.route("/editar_categoria/<int:id_categoria>", methods=["POST"])
def editar_categoria(id_categoria):
    nombre_categoria = request.form["nombre_categoria"]

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE categorias SET nombre_categoria=%s WHERE id_categoria=%s",
        (nombre_categoria, id_categoria)
    )
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("registrar_empleado"))


# ===================== AGREGAR PRODUCTO =====================
@app.route("/agregar_producto", methods=["POST"])
def agregar_producto():
    nombre = request.form["nombre"]
    precio = request.form["precio"]
    descripcion = request.form["descripcion"]
    id_categoria = request.form["id_categoria"]

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO productos_empleados (nombre, precio, descripcion, id_categoria)
        VALUES (%s, %s, %s, %s)
    """, (nombre, precio, descripcion, id_categoria))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("registrar_empleado"))


# ===================== EDITAR PRODUCTO =====================
@app.route("/editar_producto/<int:id_producto>", methods=["POST"])
def editar_producto(id_producto):
    nombre = request.form["nombre"]
    precio = request.form["precio"]
    descripcion = request.form["descripcion"]
    id_categoria = request.form["id_categoria"]

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE productos_empleados
        SET nombre=%s, precio=%s, descripcion=%s, id_categoria=%s
        WHERE id_producto=%s
    """, (nombre, precio, descripcion, id_categoria, id_producto))
    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("registrar_empleado"))


# ===================== ELIMINAR CATEGORÍA =====================
@app.route("/eliminar_categoria/<int:id_categoria>")
def eliminar_categoria(id_categoria):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM categorias WHERE id_categoria = %s", (id_categoria,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for("registrar_empleado"))


# ===================== ELIMINAR PRODUCTO =====================
@app.route("/eliminar_producto/<int:id_producto>")
def eliminar_producto(id_producto):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM productos_empleados WHERE id_producto = %s", (id_producto,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for("registrar_empleado"))


# ===================== CALCULADORA =====================
@app.route("/calculadora")
def calculadora():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()

    cursor.execute("SELECT * FROM productos_empleados")
    productos = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("calculadora.html", categorias=categorias, productos=productos)


# ===================== MESAS Y ORDENES =====================
@app.route('/mesas_empleado')
def mesas_empleado():
    return render_template('mesas_empleado.html')


@app.route('/orden/<int:mesa_id>', methods=['GET', 'POST'])
def orden_mesa(mesa_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        productos_seleccionados = request.form.getlist('producto')
        total = request.form.get('total', 0)
        tipo_entrega = request.form.get('tipo_entrega', 'restaurante')
        telefono = request.form.get('telefono', '')
        metodo_pago = request.form.get('metodo_pago', 'efectivo')

        cursor2 = conn.cursor()
        for producto_id in productos_seleccionados:
            cursor2.execute("""
                INSERT INTO pedidos (tipo_entrega, cod_mesa, fecha_pedi, hora_pedi, metodo_pago, telefono, total, estado, cod_usuario)
                VALUES (%s, %s, CURDATE(), CURTIME(), %s, %s, %s, %s, %s)
            """, (tipo_entrega, mesa_id, metodo_pago, telefono, total, 'pendiente', 0))

        conn.commit()
        cursor2.close()
        cursor.close()
        conn.close()
        return redirect(url_for('mesas_empleado'))

    cursor.execute("SELECT * FROM categorias")
    categorias = cursor.fetchall()

    cursor.execute("SELECT * FROM productos_empleados")
    productos = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('calculadora.html', mesa=mesa_id, categorias=categorias, productos=productos)


# ===================== RESERVAS =====================
@app.route("/reservas_empleado")
def reservas_empleado():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM reservas ORDER BY fecha, hora")
    reservas = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template("reservas_empleado.html", reservas=reservas)


@app.route("/agregar_reserva", methods=["POST"])
def agregar_reserva():
    fecha = request.form["fecha"]
    hora = request.form["hora"]
    cant_personas = request.form["cant_personas"]
    estado = request.form.get("estado", "disponible")
    cod_mesa = request.form["cod_mesa"]
    telefono = request.form["telefono"]
    id_usuario = request.form["id_usuario"]

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO reservas (fecha, hora, cant_personas, estado, cod_mesa, telefono, id_usuario)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (fecha, hora, cant_personas, estado, cod_mesa, telefono, id_usuario))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for("reservas_empleado"))


@app.route("/eliminar_reserva/<int:id_reservas>")
def eliminar_reserva(id_reservas):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM reservas WHERE id_reservas = %s", (id_reservas,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for("reservas_empleado"))


@app.route('/editar_reserva/<int:id>', methods=['POST'])
def editar_reserva(id):
    fecha = request.form.get("fecha")
    hora = request.form.get("hora")
    cant_personas = request.form.get("cant_personas")
    estado = request.form.get("estado")
    cod_mesa = request.form.get("cod_mesa")
    telefono = request.form.get("telefono")
    id_usuario = request.form.get("id_usuario")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE reservas 
        SET fecha=%s, hora=%s, cant_personas=%s, estado=%s, cod_mesa=%s, telefono=%s, id_usuario=%s
        WHERE id_reserva=%s
    """, (fecha, hora, cant_personas, estado, cod_mesa, telefono, id_usuario, id))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('reservas_empleado'))


@app.route('/cambiar_estado/<int:id_reservas>')
def cambiar_estado(id_reservas):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT estado FROM reservas WHERE id_reservas=%s", (id_reservas,))
    result = cursor.fetchone()
    if result:
        estado_actual = result[0]
        nuevo_estado = "no disponible" if estado_actual == "disponible" else "disponible"
        cursor.execute("UPDATE reservas SET estado=%s WHERE id_reservas=%s", (nuevo_estado, id_reservas))
        conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for('reservas_empleado'))


# ===================== ORDENES REGISTRADAS =====================
@app.route('/ordenes_empleado')
def ordenes_empleado():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM pedidos ORDER BY fecha_pedi, hora_pedi")
    ordenes = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template("ordenes_empleado.html", ordenes=ordenes)


@app.route("/eliminar_orden/<int:id_pedido>")
def eliminar_orden(id_pedido):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM pedidos WHERE id_pedido = %s", (id_pedido,))
    conn.commit()
    cursor.close()
    conn.close()
    return redirect(url_for("ordenes_empleado"))


# ===================== API JSON =====================
@app.route("/get_categorias")
def get_categorias():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id_categoria, nombre_categoria FROM categorias")
    categorias = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(categorias)


@app.route("/get_productos/<int:id_categoria>")
def get_productos(id_categoria):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT id_producto, nombre, precio
        FROM productos_empleados
        WHERE id_categoria = %s
    """, (id_categoria,))
    productos = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(productos)


# ===================== TEST DB =====================
@app.route("/testdb")
def testdb():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SHOW TABLES;")
        tables = cursor.fetchall()
        cursor.close()
        conn.close()
        return f"✅ Conectado. Tablas: {tables}"
    except Exception as e:
        return f"❌ Error conectando a MySQL: {str(e)}"


# ===================== MAIN =====================
if __name__ == "__main__":
    app.run(debug=True)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)

